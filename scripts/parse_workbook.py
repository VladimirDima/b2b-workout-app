#!/usr/bin/env python3
"""Parse Bony to Beastly Workouts xlsx into workoutData.json."""
import json
import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "Bony to Beastly Workouts 3.1.xlsx"
OUT = ROOT / "src/data/workoutData.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)

video_links = {}
for row in wb["LINKS"].iter_rows(min_row=8, values_only=True):
    name, url = row[2], row[3]
    if name and url:
        video_links[name.strip()] = url.strip()


def normalize(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


link_lookup = {normalize(k): v for k, v in video_links.items()}
ALIASES = {
    "push-ups": "push-up",
    "pit pulls": "pit pull",
    "half-kneeling pallof press against wall": "1/2 kneeling palloff press on wall",
    "hand walk-out": "hand walkout",
    "bodyweight glute bridge": "bodyweight hip thrust",
    "bent-over row": "bent-over barbell row",
    "overhead tricep extensions": "overhead tricep extension",
    "overhead triceps extensions": "overhead tricep extension",
    "underhand lat pulldowns": "underhand lat pulldown",
    "romanian deadlift": "dumbbell romanian deadlift",
}


def get_video(name: Optional[str]) -> Optional[str]:
    if not name:
        return None
    n = normalize(name)
    if n in ALIASES:
        n = ALIASES[n]
    if n in link_lookup:
        return link_lookup[n]
    for key, url in link_lookup.items():
        if n.rstrip("s") == key.rstrip("s"):
            return url
    return None


MUSCLE_MAP: Dict[str, List[str]] = {
    "Dumbbell Sumo Deadlift": ["Glutes", "Hamstrings", "Quads", "Back"],
    "Push-Ups": ["Chest", "Triceps", "Shoulders", "Core"],
    "Dumbbell Bench Press": ["Chest", "Triceps", "Shoulders"],
    "1-Arm Dumbbell Row": ["Lats", "Back", "Biceps"],
    "Dumbbell Chest Fly": ["Chest"],
    "Front Plank": ["Core", "Abs"],
    "Pit Pulls": ["Upper Back", "Rear Delts", "Traps"],
    "Overhead Tricep Extensions": ["Triceps"],
    "Overhead Triceps Extensions": ["Triceps"],
    "Goblet Squat": ["Quads", "Glutes", "Core"],
    "Underhand Lat Pulldowns": ["Lats", "Biceps"],
    "Barbell Romanian Deadlift": ["Hamstrings", "Glutes", "Back"],
    "Barbell Back Squat": ["Quads", "Glutes", "Core"],
    "Barbell Bench Press": ["Chest", "Triceps", "Shoulders"],
    "Barbell Overhead Press": ["Shoulders", "Triceps", "Core"],
    "Chin-Up": ["Lats", "Biceps", "Back"],
    "Weighted Chin-Up": ["Lats", "Biceps", "Back"],
    "Bent-Over Barbell Row": ["Lats", "Back", "Biceps"],
    "Dumbbell Romanian Deadlift": ["Hamstrings", "Glutes", "Back"],
    "Lateral Raise": ["Shoulders"],
    "Dumbbell Biceps Curl": ["Biceps"],
    "Cable Triceps Pushdown": ["Triceps"],
    "Front Squat": ["Quads", "Glutes", "Core"],
    "Barbell Front Squat": ["Quads", "Glutes", "Core"],
    "Incline Dumbbell Bench Press": ["Upper Chest", "Triceps", "Shoulders"],
    "Dumbbell Pullover": ["Chest", "Lats"],
    "Reverse Fly": ["Rear Delts", "Upper Back"],
    "Dumbbell Split Squat": ["Quads", "Glutes"],
    "Bodyweight Hip Thrust": ["Glutes", "Hamstrings"],
    "Swiss Ball Leg Curl": ["Hamstrings", "Glutes"],
    "Ab Wheel Rollout": ["Core", "Abs"],
    "Hand Walk-Out": ["Core", "Shoulders"],
    "Side Plank": ["Core", "Obliques"],
    "Half-Kneeling Pallof Press Against Wall": ["Core", "Obliques"],
    "Mountain Climbers": ["Core", "Hip Flexors"],
    "Bodyweight Glute Bridge": ["Glutes", "Hamstrings"],
    "Side Lying Leg Lift": ["Glutes", "Hip Abductors"],
    "1-Arm Farmer Carry": ["Core", "Traps", "Forearms"],
    "Bent-Over Row": ["Lats", "Back", "Biceps"],
    "Romanian Deadlift": ["Hamstrings", "Glutes", "Back"],
    "Neutral-Grip Seated Cable Row": ["Lats", "Back", "Biceps"],
    "Lat Pulldown": ["Lats", "Biceps"],
    "Standing Landmine Press": ["Shoulders", "Core"],
    "Dumbbell Hammer Curl": ["Biceps", "Forearms"],
    "EZ-Bar Biceps Curl": ["Biceps"],
    "Dumbbell Skull Crusher": ["Triceps"],
    "Close-Grip Push-Up": ["Chest", "Triceps"],
    "Deficit Push-Up": ["Chest", "Triceps", "Shoulders"],
    "TYIs on Incline Bench": ["Rear Delts", "Upper Back", "Traps"],
    "Standing External Rotation": ["Rotator Cuff", "Shoulders"],
    "Kettlebell Swing": ["Glutes", "Hamstrings", "Core"],
    "Barbell Deadlift": ["Glutes", "Hamstrings", "Back", "Traps"],
    "Chest-Supported Dumbbell Row": ["Lats", "Back", "Biceps"],
    "Overhand Seated Cable Row": ["Lats", "Back", "Biceps"],
    "Underhand Cable Seated Row": ["Lats", "Biceps"],
    "Barbell Split Squat": ["Quads", "Glutes"],
    "Dumbbell Farmer Carry": ["Core", "Traps", "Forearms"],
    "Cross-Body Carry": ["Core", "Obliques", "Traps"],
    "Suitcase Romanian Deadlift": ["Hamstrings", "Glutes", "Core"],
    "Offset Goblet Squat": ["Quads", "Glutes", "Core"],
    "Raised Deadlift": ["Hamstrings", "Glutes", "Back"],
    "Swiss Ball Rollout": ["Core", "Abs"],
    "Weighted Crunches": ["Abs", "Core"],
    "Cable Crunches": ["Abs", "Core"],
    "Hanging Knee Raise": ["Abs", "Core"],
    "Nordic Ham Curl": ["Hamstrings"],
    "Machine Leg Curl": ["Hamstrings"],
    "Leg Press": ["Quads", "Glutes"],
    "Machine Chest Press": ["Chest", "Triceps"],
    "Machine Fly/Pec Deck": ["Chest"],
    "Machine Reverse Fly": ["Rear Delts", "Upper Back"],
    "Machine Biceps Curl": ["Biceps"],
    "T-Bar Machine Row": ["Lats", "Back"],
    "1-Arm Incline Dumbbell Bench": ["Upper Chest", "Triceps"],
    "Half-Kneeling Landmine Press": ["Shoulders", "Core"],
    "Half-Kneeling Dumbbell Press": ["Shoulders", "Triceps"],
    "Alternating Dumbbell Bench Press": ["Chest", "Triceps"],
    "Double-Dumbbell Front Squat": ["Quads", "Glutes", "Core"],
    "Dumbbell Step-Up": ["Quads", "Glutes"],
    "Bodyweight Step-Up": ["Quads", "Glutes"],
    "Bodyweight Split Squat": ["Quads", "Glutes"],
    "Cable Rope Curl": ["Biceps"],
    "Incline Barbell Bench Press": ["Upper Chest", "Triceps", "Shoulders"],
    "Barbell Biceps Curl": ["Biceps"],
    "Dumbbell Swing": ["Glutes", "Hamstrings", "Core"],
    "Weighted Push-Up": ["Chest", "Triceps", "Shoulders"],
    "Weighted Side Plank": ["Core", "Obliques"],
    "Below-the-Knee Rack Pull": ["Back", "Glutes", "Traps"],
    "Barbell Curl": ["Biceps"],
    "Close-Grip Barbell Bench Press": ["Chest", "Triceps"],
    "Dumbbell Racked Carry": ["Core", "Traps"],
}


def get_muscles(name: str) -> List[str]:
    if name in MUSCLE_MAP:
        return MUSCLE_MAP[name]
    n = name.lower()
    muscles = set()
    rules = [
        (["squat", "lunge", "step-up", "leg press"], ["Quads", "Glutes"]),
        (["deadlift", "rdl", "romanian", "rack pull"], ["Hamstrings", "Glutes", "Back"]),
        (["bench press", "push-up", "push up", "chest fly", "pec"], ["Chest", "Triceps"]),
        (["row", "pulldown", "chin", "pull-up"], ["Lats", "Back", "Biceps"]),
        (["curl"], ["Biceps"]),
        (["tricep", "skull crusher", "pushdown", "extension"], ["Triceps"]),
        (["overhead press", "landmine press", "lateral raise"], ["Shoulders"]),
        (["plank", "crunch", "rollout", "ab wheel", "pallof"], ["Core", "Abs"]),
        (["fly", "reverse fly", "tyi"], ["Rear Delts", "Upper Back"]),
        (["carry", "farmer"], ["Core", "Traps"]),
        (["hip thrust", "glute bridge", "swing"], ["Glutes", "Hamstrings"]),
        (["leg curl", "nordic"], ["Hamstrings"]),
        (["shrug"], ["Traps"]),
        (["pit pull"], ["Upper Back", "Traps"]),
        (["rotation"], ["Rotator Cuff"]),
        (["mountain climber", "leg lift"], ["Core", "Hip Flexors"]),
    ]
    for keywords, ms in rules:
        if any(k in n for k in keywords):
            muscles.update(ms)
    return sorted(muscles) if muscles else ["Full Body"]


def fmt_val(v):
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        return int(v)
    if isinstance(v, str):
        return v.strip()
    return str(v)


def is_metadata(text: str) -> bool:
    if not text:
        return False
    t = text.upper().strip()
    return t in ("EXERCISE", "VIDEO TUTORIAL")


def is_note_only(text: str) -> bool:
    prefixes = ("REST:", "RIR:", "AMRAP:", "IT'S OKAY", "VIDEO TUTORIAL", "(")
    t = text.strip()
    if not t:
        return True
    if any(t.upper().startswith(p) for p in prefixes if p != "("):
        return True
    if t.startswith("("):
        return True
    if re.search(r"stop\s+\d+\s+reps?\s+shy", t, re.I):
        return True
    if "REPS IN RESERVE" in t.upper():
        return True
    return False


def extract_rir_number(text: str):
    m = re.search(r"RIR:\s*(\d+)", text, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"\((\d+)\s*Reps?\s+in\s+Reserve\)", text, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)\s*Reps?\s+in\s+Reserve", text, re.I)
    if m:
        return int(m.group(1))
    m = re.search(r"stop\s+(\d+)\s+reps?\s+shy", text, re.I)
    if m:
        return int(m.group(1))
    return None


def classify_note(text: str) -> dict:
    if not text:
        return {}
    t = text.strip()
    u = t.upper()
    if u.startswith("RIR:") or "REPS IN RESERVE" in u or re.search(r"stop\s+\d+\s+reps?\s+shy", t, re.I):
        rir = extract_rir_number(t)
        if rir is not None:
            return {"rir": rir, "rirNote": t}
    if u.startswith("REST:"):
        return {"rest": t}
    if u.startswith("VIDEO TUTORIAL"):
        return {}
    return {"notes": t}


def finalize_progression(progression: list) -> list:
    exercise_rir = None
    exercise_rir_note = None
    for entry in progression:
        if entry.get("rir") is not None:
            exercise_rir = entry["rir"]
            exercise_rir_note = entry.get("rirNote")
            break

    current_rest = None
    for entry in progression:
        if exercise_rir is not None:
            entry["rir"] = exercise_rir
            if exercise_rir_note:
                entry["rirNote"] = exercise_rir_note

        if entry.get("rest"):
            current_rest = entry["rest"]
        elif current_rest:
            entry["rest"] = current_rest
    return progression


def has_prog(week, sets, reps) -> bool:
    if week is None or sets is None or reps is None:
        return False
    if isinstance(week, str) and week.strip().upper() in ("WEEK", "EXERCISE"):
        return False
    return True


def parse_phase_sheet(ws, phase_id: str, phase_name: str) -> dict:
    days, current_day, current_block, current_exercise = [], None, None, None
    for row in ws.iter_rows(min_row=1, values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        day_col = row[5] if len(row) > 5 else None
        week = row[4] if len(row) > 4 else None
        sets = row[5] if len(row) > 5 else None
        reps = row[6] if len(row) > 6 else None

        if day_col and isinstance(day_col, str) and re.match(
            r"DAY\s+(ONE|TWO|THREE|FOUR|FIVE|SIX)", day_col.upper()
        ):
            num = {"ONE": 1, "TWO": 2, "THREE": 3, "FOUR": 4, "FIVE": 5, "SIX": 6}[
                day_col.split()[-1].upper()
            ]
            current_day = {"id": f"day-{num}", "name": day_col.strip(), "blocks": []}
            days.append(current_day)
            current_block = None
            current_exercise = None
            continue

        if b and isinstance(b, str):
            bu = b.upper()
            if any(x in bu for x in ["SUPERSET", "GIANT SET", "OPTIONAL"]) and current_day:
                current_block = {
                    "id": f"block-{len(current_day['blocks']) + 1}",
                    "name": b.strip(),
                    "exercises": [],
                }
                current_day["blocks"].append(current_block)
                current_exercise = None
                continue

        if not current_day or not current_block:
            continue

        text = str(c).strip() if c else ""
        if is_metadata(text) and not has_prog(week, sets, reps):
            continue

        if has_prog(week, sets, reps):
            note_fields = classify_note(text) if is_note_only(text) else {}
            entry = {
                "week": fmt_val(week),
                "sets": fmt_val(sets),
                "reps": fmt_val(reps),
                "notes": note_fields.pop("notes", "") if note_fields else "",
                **note_fields,
            }
            if is_note_only(text) or not text:
                if current_exercise:
                    current_exercise["progression"].append(entry)
            else:
                current_exercise = {
                    "name": text,
                    "videoUrl": get_video(text),
                    "muscleGroups": get_muscles(text),
                    "progression": [entry],
                }
                current_block["exercises"].append(current_exercise)
        elif text and not is_note_only(text):
            current_exercise = {
                "name": text,
                "videoUrl": get_video(text),
                "muscleGroups": get_muscles(text),
                "progression": [],
            }
            current_block["exercises"].append(current_exercise)

    for day in days:
        for block in day["blocks"]:
            for exercise in block["exercises"]:
                exercise["progression"] = finalize_progression(exercise["progression"])

    return {"id": phase_id, "name": phase_name, "days": days}


def parse_warmups(ws) -> List[dict]:
    sections, current = [], None
    for row in ws.iter_rows(min_row=1, values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        g = row[6] if len(row) > 6 else None
        header = b or c
        if header and isinstance(header, str) and any(
            x in header.upper() for x in ["CORE FOUR", "HIP PREP", "BIG FIVE"]
        ):
            current = {"name": header.strip(), "exercises": []}
            sections.append(current)
        elif current and c and isinstance(c, str) and g and isinstance(g, str):
            if not any(x in c.upper() for x in ["RECOMMENDED", "OPTIONAL", "USE 20"]):
                current["exercises"].append(
                    {
                        "name": c.strip(),
                        "instructions": g.strip(),
                        "videoUrl": get_video(c.strip()),
                        "muscleGroups": get_muscles(c.strip()),
                    }
                )
    return sections


warmups = parse_warmups(wb["WARM-UPS"])
phases = []
for pid, sheet, title in [
    ("phase-0", "PHASE 0", "Phase Zero — The Newbie"),
    ("phase-1", "PHASE 1", "Phase One — The Brute"),
    ("phase-2", "PHASE 2", "Phase Two — The Iron Gut"),
    ("phase-3", "PHASE 3", "Phase Three — The Mass Hound"),
    ("phase-4", "PHASE 4", "Phase Four — The Powerbuilder"),
]:
    phases.append(parse_phase_sheet(wb[sheet], pid, title))

data = {
    "program": {
        "name": "Bony to Beastly Workouts",
        "version": "3.1",
        "warmups": warmups,
        "phases": phases,
        "videoLinks": video_links,
    }
}

OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2)

print(f"Wrote {OUT}")
